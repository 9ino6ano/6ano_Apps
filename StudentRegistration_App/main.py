from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox

import openpyxl
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
from openpyxl import Workbook
import pathlib

background="#06283D"
frame_bg="#EDEDED"
frame_fg="#06283D"

root=Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "D_O_B"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Mother's Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"

    file.save('Student_data.xlsx')

#gender select function
def selection():
    global Gender
    value=radio.get()
    if value==1:
        Gender="Male"
        print(Gender)
    else:
        Gender="Female"
        print(Gender)


def Upload():
    global file_name
    global img
    file_name=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetypes=(("JPG File","*.jpg"),
                                                                                                     ("PNG File","*.png"),
                                                                                                     ("All files","*.txt")))
    img = (Image.open(file_name))
    resized_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#Registration No -> now each time we have to enter a registration no
#Lets design automatic registration no entry system.
def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    #This will check data of the last row and add 1 to the registration no. & also if there is not any number it will give 1
    try:
        reg_number.set(max_row_value+1)

    except:
        reg_number.set(1)


def Save():
    R1=reg_no.get()
    N1=Name.get()
    C1=st_class.get()
    try:
        G1=Gender #This will show only if you havent selected a gender
    except:
        messagebox.showerror("Error","Please specify which gender are you!!")

    D2=date_of_birth.get()
    D1=date_today.get()
    Rel=Religion.get()
    S1=Skills.get()
    father_name=f_name.get()
    mother_name=m_name.get()
    F1=f_occupation.get()
    m1=m_occupation.get()

    """
    [This will test whether the input fields are doing exactly what needs to be done
    e.g capturing the user details on the entry form]
    print(R1)
    print(N1)
    print(C1)
    print(G1)
    print(D2)
    print(D1)
    print(Rel)
    print(S1)
    print(father_name)
    print(mother_name)
    print(F1)
    print(m1)
    """
    if N1=="" or C1=="Select Class" or D2=="" or Rel=="" or S1=="" or father_name=="" or mother_name == "" or F1=="" or m1=="":
        messagebox.showerror("Data Missing","Please input your information on all the entry fields provided.")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row+1,value=N1)
        sheet.cell(column=3, row=sheet.max_row + 1, value=C1)
        sheet.cell(column=4, row=sheet.max_row + 1, value=G1)
        sheet.cell(column=5, row=sheet.max_row + 1, value=D1)
        sheet.cell(column=6, row=sheet.max_row + 1, value=D2)
        sheet.cell(column=7, row=sheet.max_row + 1, value=Rel)
        sheet.cell(column=8, row=sheet.max_row + 1, value=S1)
        sheet.cell(column=9, row=sheet.max_row + 1, value=father_name)
        sheet.cell(column=10, row=sheet.max_row + 1, value=mother_name)
        sheet.cell(column=11, row=sheet.max_row + 1, value=F1)
        sheet.cell(column=12, row=sheet.max_row + 1, value=m1)

        file.save(r'Student_data.xlsx')
        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("Information","The profile picture is unavailable at this moment.")

        messagebox.showinfo("Info","Successfully Added Data")

        Reset() # This will clear the controls on the form/screen

        registration_no() #This will check if there's a reg_number, if so it will issue out a new 1

##################################
def search():
    text = Search.get() #taking input from the entry box
    Reset() #To clear all the data already available in the entry box and others
    save_btn.config(state='disable') #after clicking on the search, save button will disable so that no can click on it

    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            #print(str(name))
            reg_no_pos=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid Registration Number!!!!")


#reg_no_pos -> showing like A2,A3,A4,A5 although the reg_number show A2,3,4,....
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    """
    print(x1)
    print(x2)
    print(x3)
    print(x4)
    print(x5)
    print(x6)
    print(x7)
    print(x8)
    print(x9)
    print(x10)
    print(x11)
    print(x12)
    """
    reg_number.set(x1)
    Name.set(x2)
    st_class.set(x3)
    if x4=='Female':
        R2.select()
    else:
        R1.select()
    date_of_birth.set(x5)
    date_today.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    f_name.set(x9)
    m_name.set(x10)
    f_occupation.set(x11)
    m_occupation.set(x12)

    #We are doing this to take the same image as the registration number
    img = (Image.open("Student Images/"+str(x1)+".jpg"))
    resized_img = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_img)
    lbl.config(image=photo2)
    lbl.image=photo2
#############
#The Update button
def Update():
    print("This is working")
    R1=reg_no.get()
    N1=Name.get()
    C1=st_class.get()
    selection()
    G1=Gender
    D2=date_of_birth.get()
    D1=date_today.get()
    Rel=Religion.get()
    S1=Skills.get()
    father_name=f_name.get()
    mother_name=m_name.get()
    F1=f_occupation.get()
    m1=m_occupation.get()

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name=row[0]
            print(str(name))
            reg_no_pos=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

    #sheet.cell(column=1,row=int(reg_number),value=R1) -> Nobody must be able to update registration umber it must remain unique
    sheet.cell(column=1, row=int(reg_number), value=N1)
    sheet.cell(column=1, row=int(reg_number), value=C1)
    sheet.cell(column=1, row=int(reg_number), value=G1)
    sheet.cell(column=1, row=int(reg_number), value=D1)
    sheet.cell(column=1, row=int(reg_number), value=D2)
    sheet.cell(column=1, row=int(reg_number), value=Rel)
    sheet.cell(column=1, row=int(reg_number), value=S1)
    sheet.cell(column=1, row=int(reg_number), value=father_name)
    sheet.cell(column=1, row=int(reg_number), value=mother_name)
    sheet.cell(column=1, row=int(reg_number), value=F1)
    sheet.cell(column=1, row=int(reg_number), value=m1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Image/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Updated","Information has been updated successfully.")
    Reset()


def Reset():
    global img
    Name.set("")
    date_of_birth.set("")
    Religion.set("")
    Skills.set("")
    f_name.set("")
    m_name.set("")
    f_occupation.set("")
    m_occupation.set("")
    st_class.set("Select a class")
    reg_no.setvar("")

    registration_no()
    save_btn.config(state="normal")
    img1=PhotoImage(file='images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

def Exit():
    messagebox.showinfo("Exit Application","Are you sure you want to exit")
    root.destroy()



#Top frames
Label(root,text="Email: panikiesvanstouter@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="Student Registration No: ",width=10,height=2,bg="#c36464",fg="#fff",font=('arial 20 bold')).pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
image_icon3=PhotoImage(file="Images/search.png")
search_btn=Button(root,text="Search",compound=LEFT,image=image_icon3,width=123,bg="#68ddfa",font="font 13 bold",command=search)
search_btn.place(x=1060,y=66)

image_icon4=PhotoImage(file="Images/Layer 4.png")
update_btn=Button(root,image=image_icon4,width=123,bg="#C36464",command=Update)
update_btn.place(x=110,y=64)

#Registration and Date
Label(root,text="Registration No:" ,font="arial 13",fg=frame_bg,bg=background).place(x=30,y=150)
Label(root,text="Date:" ,font="arial 13",fg=frame_bg,bg=background).place(x=500,y=150)

reg_number = IntVar()
date_today = StringVar()

reg_no=Entry(root,textvariable=reg_number,width=15,font="arial 10")
reg_no.place(x=160,y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")

date_entry = Entry(root,textvariable=date_today,width=15,font="arial 10")
date_entry.place(x=550,y=150)
date_today.set(d1)

#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=frame_bg,fg=frame_fg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=50,y=50)
Label(obj,text="Date Of Birth: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=50,y=100)
Label(obj,text="Gender: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=50,y=150)
Label(obj,text="Class: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=500,y=50)
Label(obj,text="Religion: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=500,y=100)
Label(obj,text="Skills: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=500,y=150)

Name=StringVar()
name_Entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_Entry.place(x=160,y=50)

Gender=StringVar()

radio=IntVar()
R1 = Radiobutton(obj,text="Male",variable=radio,value=1,bg=frame_bg,fg=frame_fg,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj,text="Female",variable=radio,value=2,bg=frame_bg,fg=frame_fg,command=selection)
R2.place(x=200,y=150)

date_of_birth=StringVar()
dob_Entry=Entry(obj,textvariable=date_of_birth,width=20,font="arial 10")
dob_Entry.place(x=160,y=100)

st_class = Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12','Other'],font="Roboto 10",width=17,state='r')
st_class.place(x=630,y=50)
st_class.set("Select Class")

Religion=StringVar()
Religion_Entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")
Religion_Entry.place(x=630,y=100)

Skills=StringVar()
skills_Entry=Entry(obj,textvariable=Skills,width=20,font="arial 10")
skills_Entry.place(x=630,y=150)


#Parent details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=frame_bg,fg=frame_fg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)
Label(obj2,text="Father's Name: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=30,y=100)
f_name=StringVar()
f_entry=Entry(obj2,textvariable=f_name,width=20,font="arial 10")
f_entry.place(x=160,y=50)
f_occupation=StringVar()
f_entry=Entry(obj2,textvariable=f_occupation,width=20,font="arial 10")
f_entry.place(x=160,y=100)

Label(obj2,text="Mother's Name: ",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=frame_bg,fg=frame_fg).place(x=500,y=100)
m_name=StringVar()
m_entry=Entry(obj2,textvariable=m_name,width=20,font="arial 10")
m_entry.place(x=630,y=50)
m_occupation=StringVar()
m_entry=Entry(obj2,textvariable=m_occupation,width=20,font="arial 10")
m_entry.place(x=630,y=100)

#Image
f = Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)
img=PhotoImage(file="Images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#Button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Upload).place(x=1000,y=370)

save_btn=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
save_btn.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="yellow",command=Reset).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="red",command=Exit).place(x=1000,y=610)





root.mainloop()